"""One-time/re-runnable import for the PSR (Periodic Status Report) Tracking
Log workbook. Each sheet maps to a category:

  RPOSD                          -> "RPOSD"
  State of California NRA        -> "State of California NRA"
  Conservancy (RMC,BHUWC)        -> "Conservancy (RMC,BHUWC)"
  Cooling Amenities and Fairfax  -> "Cooling Amenities and Fairfax"
  Other                          -> "Other"

Column layout differs per sheet (see COLUMN_MAP below). Footer/annotation
rows (starting with "*", or a "Count" summary row) are skipped.

Safe to re-run: projects are matched by (project_name, category) and
updated in place. Due dates are matched by (project_id, due_date) so
re-running doesn't create duplicates; existing due dates are updated
in place (submitted/submitted_date) rather than duplicated.

Usage:
    python scripts/import_psr_log.py path/to/file.xlsx
"""
import sys
from datetime import date, datetime
from pathlib import Path

BACKEND_DIR = Path(__file__).resolve().parents[1] / "backend"
sys.path.insert(0, str(BACKEND_DIR))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from app.models.models import PSRDueDate, PSRProject  # noqa: E402

# column index (0-based) -> field, per sheet
COLUMN_MAP = {
    "RPOSD": {
        "project_name": 0,
        "ad_number": 1,
        "psr_due_date": 4,
        "performance_end_date": 5,
        "grant_manager": 6,
        "submission_date": 7,
    },
    "State of California NRA": {
        "project_name": 0,
        "funding_source": 1,
        "psr_due_date": 2,
        "submission_date": 3,
    },
    "Conservancy (RMC,BHUWC)": {
        "project_name": 0,
        "psr_due_date": 1,
        "district": 2,
        "grant_manager": 3,
        "submission_date": 4,
    },
    "Cooling Amenities and Fairfax": {
        "project_name": 0,
        "psr_due_date": 1,
        "district": 2,
        "grant_manager": 3,
        "submitted_yn": 4,
    },
    "Other": {
        "project_name": 0,
        "grantor": 1,
        "funding_source": 2,
        "psr_due_date": 3,
        "submission_date": 4,
    },
}


def cell_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def cell_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def cell_int(value) -> int | None:
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None


def is_footer_row(project_name: str | None) -> bool:
    if not project_name:
        return True
    s = project_name.strip()
    if not s:
        return True
    if s.startswith("*"):
        return True
    if s.lower() == "count":
        return True
    return False


def parse_sheet(ws, sheet_name: str) -> list[dict]:
    cols = COLUMN_MAP[sheet_name]

    def get(row, key):
        idx = cols.get(key)
        return row[idx].value if idx is not None else None

    def get_cell(row, key):
        idx = cols.get(key)
        return row[idx] if idx is not None else None

    records = []
    for row in ws.iter_rows(min_row=2):
        project_name = cell_str(get(row, "project_name"))
        if is_footer_row(project_name):
            continue

        due_date = cell_date(get(row, "psr_due_date"))
        submission_date = cell_date(get(row, "submission_date"))
        submitted_yn = cell_str(get(row, "submitted_yn"))

        # AD Number cells (RPOSD) carry a real hyperlink to the grantee portal
        # record for that grant; other sheets have no cell hyperlinks at all.
        link = None
        ad_number_cell = get_cell(row, "ad_number")
        if ad_number_cell is not None and ad_number_cell.hyperlink:
            link = ad_number_cell.hyperlink.target

        submitted = False
        if submitted_yn is not None:
            submitted = submitted_yn.strip().upper() == "Y"
        elif submission_date is not None:
            submitted = True

        # Sheets other than "Other" have no distinct Grantor column -- the
        # category itself is the grantor for those.
        grantor = cell_str(get(row, "grantor")) or (sheet_name if "grantor" not in cols else None)

        records.append(
            {
                "project_name": project_name,
                "category": sheet_name,
                "grantor": grantor,
                "funding_source": cell_str(get(row, "funding_source")),
                "ad_number": cell_str(get(row, "ad_number")),
                "district": cell_int(get(row, "district")),
                "grant_manager": cell_str(get(row, "grant_manager")),
                "performance_end_date": cell_date(get(row, "performance_end_date")),
                "link": link,
                "due_date": due_date,
                "submitted": submitted,
                "submitted_date": submission_date,
            }
        )
    return records


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/import_psr_log.py path/to/file.xlsx", file=sys.stderr)
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    if not xlsx_path.exists():
        print(f"File not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    db = SessionLocal()
    try:
        existing_by_key = {
            (p.project_name.strip().lower(), p.category): p for p in db.scalars(select(PSRProject)).all()
        }

        projects_created = 0
        projects_updated = 0
        due_dates_created = 0
        due_dates_updated = 0

        for sheet_name in COLUMN_MAP:
            if sheet_name not in wb.sheetnames:
                print(f"Warning: sheet {sheet_name!r} not found in workbook, skipping", file=sys.stderr)
                continue
            records = parse_sheet(wb[sheet_name], sheet_name)

            for rec in records:
                key = (rec["project_name"].strip().lower(), rec["category"])
                project = existing_by_key.get(key)
                if project is None:
                    project = PSRProject(
                        project_name=rec["project_name"],
                        category=rec["category"],
                        grantor=rec["grantor"],
                        funding_source=rec["funding_source"],
                        ad_number=rec["ad_number"],
                        district=rec["district"],
                        grant_manager=rec["grant_manager"],
                        performance_end_date=rec["performance_end_date"],
                        link=rec["link"],
                    )
                    db.add(project)
                    db.flush()
                    existing_by_key[key] = project
                    projects_created += 1
                else:
                    project.grantor = rec["grantor"] or project.grantor
                    project.funding_source = rec["funding_source"] or project.funding_source
                    project.ad_number = rec["ad_number"] or project.ad_number
                    project.district = rec["district"] if rec["district"] is not None else project.district
                    project.grant_manager = rec["grant_manager"] or project.grant_manager
                    project.performance_end_date = rec["performance_end_date"] or project.performance_end_date
                    project.link = rec["link"] or project.link
                    projects_updated += 1

                if rec["due_date"] is not None:
                    existing_due = db.scalar(
                        select(PSRDueDate).where(
                            PSRDueDate.project_id == project.id, PSRDueDate.due_date == rec["due_date"]
                        )
                    )
                    if existing_due is None:
                        db.add(
                            PSRDueDate(
                                project_id=project.id,
                                due_date=rec["due_date"],
                                submitted=rec["submitted"],
                                submitted_date=rec["submitted_date"],
                            )
                        )
                        due_dates_created += 1
                    else:
                        existing_due.submitted = rec["submitted"]
                        existing_due.submitted_date = rec["submitted_date"]
                        due_dates_updated += 1

        db.commit()

        print(f"Projects created: {projects_created}")
        print(f"Projects updated: {projects_updated}")
        print(f"Due dates created: {due_dates_created}")
        print(f"Due dates updated: {due_dates_updated}")
    finally:
        db.close()


if __name__ == "__main__":
    main()
